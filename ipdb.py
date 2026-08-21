#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IP 归属业务域（方案一拆分）
功能：离线/在线 IP 归属查询、内外网/排除 IP 判定、终端 IP 表、IP归属分析 xlsx 生成。
依赖：common（日志/路径/回调）、updater（safe_get / get_requests_verify）、stdlib。
注意：load_config 也放本模块——ipdb 与 report 均需读取 config.ini（config 内部调用
load_terminal_ip_table），放此处可保证 report→ipdb 单向依赖、无循环 import。
"""

import os
import re
import sys
import time
import ipaddress
import configparser
from pathlib import Path
from datetime import datetime

from common import (script_dir, runtime_dir, _log, _set_progress, _find_file)
from updater import safe_get, get_requests_verify


def _load_excluded_ip_networks(config_path: str = None) -> list:
    if config_path is None:
        config_path = os.path.join(runtime_dir, 'config.ini')
    config = configparser.ConfigParser()
    config.read(config_path, encoding='utf-8')
    excluded_ips_str = config.get('network', 'excluded_ips', fallback='')
    networks = []
    for _ex_range in (r.strip() for r in excluded_ips_str.split(',') if r.strip()):
        if '-' not in _ex_range:
            try:
                networks.append(ipaddress.ip_address(_ex_range))
            except ValueError:
                pass
            continue
        parts = _ex_range.split('-')
        if len(parts) != 2:
            continue
        base, end_part = parts[0].strip(), parts[1].strip()
        # 简写展开：1.2.3.4-10 → 1.2.3.4-1.2.3.10（与 _parse_ip_range 行为一致）
        end_parts = end_part.split('.')
        if len(end_parts) == 1:
            base_parts = base.split('.')
            if len(base_parts) == 4:
                end_part = f'{base_parts[0]}.{base_parts[1]}.{base_parts[2]}.{end_parts[0]}'
            else:
                continue
        try:
            start_ip = ipaddress.ip_address(base)
            end_ip = ipaddress.ip_address(end_part)
            if int(start_ip) > int(end_ip):
                start_ip, end_ip = end_ip, start_ip
            networks.extend(ipaddress.summarize_address_range(start_ip, end_ip))
        except ValueError:
            pass
    return networks

EXCLUDED_IP_NETWORKS = _load_excluded_ip_networks()
EXCLUDED_IP_LABELS = {}  # IP -> 说明（来自外部Excel）

def is_excluded_ip(ip_str):
    try:
        ip = ipaddress.ip_address(str(ip_str).strip())
    except ValueError:
        return False
    for net in EXCLUDED_IP_NETWORKS:
        if isinstance(net, ipaddress.IPv4Address):
            if ip == net:
                return True
        else:
            if ip in net:
                return True
    return False

def load_external_excluded_ips(excel_path=None):
    """从外部Excel文件加载排除业务IP，格式：ip, 说明。

    excel_path 缺省时自动查找脚本目录下的 业务ip.xlsx（解放 config.ini 的
    excluded_ips 硬编码）。支持单 IP / 范围 / 简写（1.2.3.4-10）。
    """
    global EXCLUDED_IP_NETWORKS, EXCLUDED_IP_LABELS
    if excel_path is None:
        excel_path = _find_file('业务ip.xlsx')
    if not excel_path or not os.path.exists(excel_path):
        _log('[!] 未找到业务ip.xlsx（排除IP清单），跳过自动加载')
        return 0
    import pandas as pd
    try:
        df = pd.read_excel(excel_path)
        df.columns = df.columns.str.strip()
        # 识别IP列和说明列
        ip_col = None
        label_col = None
        for col in df.columns:
            col_lower = str(col).lower()
            if 'ip' in col_lower and ip_col is None:
                ip_col = col
            elif '说明' in str(col) or '备注' in str(col) or 'desc' in col_lower:
                label_col = col
        if ip_col is None:
            # 如果没找到带ip的列名，取第一列作为IP
            ip_col = df.columns[0]
            if label_col is None and len(df.columns) > 1:
                label_col = df.columns[1]
        count = 0
        for _, row in df.iterrows():
            ip_val = row[ip_col]
            if pd.isna(ip_val):
                continue
            ip_str = str(ip_val).strip()
            label_str = str(row[label_col]).strip() if label_col and not pd.isna(row[label_col]) else '业务IP'
            # 复用 _parse_ip_range 统一处理单IP/范围/简写
            ips = _parse_ip_range(ip_str)
            if not ips:
                continue
            start_ip = ipaddress.ip_address(ips[0])
            end_ip = ipaddress.ip_address(ips[-1])
            if len(ips) == 1:
                EXCLUDED_IP_NETWORKS.append(start_ip)
            else:
                for net in ipaddress.summarize_address_range(start_ip, end_ip):
                    EXCLUDED_IP_NETWORKS.append(net)
            for ip in ips:
                EXCLUDED_IP_LABELS[ip] = label_str
            count += 1
        _log(f'[+] 从业务IP文件加载排除IP: {count} 条 ({os.path.basename(excel_path)})')
        return count
    except Exception as e:
        _log(f'[!] 加载外部业务IP文件失败: {e}')
        return 0


# 自动加载业务IP 的幂等标志（避免多次调用重复 append）
_AUTO_EXCLUDED_LOADED = False

def _auto_load_excluded_ips():
    """自动从 业务ip.xlsx 加载排除IP（仅首次生效）"""
    global _AUTO_EXCLUDED_LOADED
    if _AUTO_EXCLUDED_LOADED:
        return
    _AUTO_EXCLUDED_LOADED = True
    load_external_excluded_ips()


def extract_zones_from_alerts(files):
    """从安全告警文件的"源区域"列提取内网区域集合（自动发现，解放 config internal_zones）。

    排除"默认区域"等无业务含义的噪声值；文件列名自动 strip 容错。
    """
    import pandas as pd
    zones = set()
    for f in files:
        path = str(f)
        try:
            df = pd.read_excel(path, usecols=['源区域'])
        except Exception:
            try:
                df = pd.read_excel(path)
            except Exception:
                continue
            df.columns = df.columns.str.strip()
            if '源区域' not in df.columns:
                continue
            df = df[['源区域']]
        for v in df['源区域'].dropna().astype(str):
            v = v.strip()
            if v and v != '默认区域':
                zones.add(v)
    if zones:
        _log(f'[+] 从告警文件提取内网区域 {len(zones)} 个: {", ".join(sorted(zones)[:8])}{"..." if len(zones) > 8 else ""}')
    return zones


def extract_geos_from_alerts(files):
    """从安全告警文件的"源地理信息"列提取本地归属地关键词（吉林-长春 → 长春）。

    命中"吉林/长春/本地"相关值才提取，避免把北京/国外等误判为本地。
    """
    import pandas as pd
    geos = set()
    hints = ('吉林', '长春', '本地')
    for f in files:
        path = str(f)
        try:
            df = pd.read_excel(path, usecols=['源地理信息'])
        except Exception:
            try:
                df = pd.read_excel(path)
            except Exception:
                continue
            df.columns = df.columns.str.strip()
            if '源地理信息' not in df.columns:
                continue
            df = df[['源地理信息']]
        for v in df['源地理信息'].dropna().astype(str):
            if any(h in v for h in hints):
                geos.add('长春')
    if geos:
        _log(f'[+] 从告警文件提取本地归属地关键词: {", ".join(sorted(geos))}')
    return geos


def load_probes_from_excel():
    """从 业务ip.xlsx 的"探针"sheet 读取探针（列：名称 | IP地址）。

    sheet 名模糊匹配（含"探针"即可，如 探针 / 探针ip段 / 探针配置）。
    无该 sheet 或无有效行时返回 []（回退 config.ini [health] probes）。
    """
    path = _find_file('业务ip.xlsx')
    if not path or not os.path.exists(path):
        return []
    try:
        import pandas as pd
        xl = pd.ExcelFile(path)
        probe_sheet = next((s for s in xl.sheet_names if '探针' in s), None)
        if not probe_sheet:
            return []
        df = pd.read_excel(path, sheet_name=probe_sheet)
        df.columns = df.columns.str.strip()
        name_col = '名称' if '名称' in df.columns else df.columns[0]
        ip_col = next((c for c in ('IP地址', 'IP', 'ip') if c in df.columns), df.columns[1] if len(df.columns) > 1 else name_col)
        rows = []
        for _, r in df.iterrows():
            name = str(r[name_col]).strip()
            ip = str(r[ip_col]).strip()
            if name and ip and ip.lower() != 'nan':
                rows.append((name, ip))
        if rows:
            _log(f'[+] 从业务ip.xlsx[{probe_sheet}]加载探针: {len(rows)} 个')
        return rows
    except Exception as e:
        _log(f'[!] 读取探针sheet失败: {e}')
        return []

def is_private_ip(ip_str):
    try:
        ip = ipaddress.ip_address(str(ip_str).strip())
        return ip.is_private or ip.is_loopback or ip.is_reserved or ip.is_link_local or ip.is_multicast
    except ValueError:
        return False

TERMINAL_IP_TABLE = None
TERMINAL_IP_TABLE_PATH = None  # 外部导入的终端IP表路径

def set_terminal_ip_table_path(path):
    """设置终端IP地址表路径并重置缓存，供GUI导入按钮调用"""
    global TERMINAL_IP_TABLE, TERMINAL_IP_TABLE_PATH
    TERMINAL_IP_TABLE_PATH = path
    TERMINAL_IP_TABLE = None  # 重置缓存，下次调用时重新加载

def _parse_ip_range(ip_str):
    """解析IP字符串，支持单IP和范围格式（完整/简写），返回IP列表。
    支持格式:
      - 单IP: '172.16.70.226'
      - 完整范围: '172.16.70.226-172.16.70.230'
      - 简写范围: '172.16.70.226-230'
    异常时返回空列表并打印警告日志。
    """
    result = []
    ip_str = str(ip_str).strip()
    if not ip_str:
        return result
    try:
        if '-' not in ip_str:
            # 单IP
            ipaddress.ip_address(ip_str)  # 校验格式
            result.append(ip_str)
            return result

        parts = ip_str.split('-')
        if len(parts) != 2:
            _log(f'[!] 跳过无效IP格式（包含多个"-"）: {ip_str}')
            return result

        start_str = parts[0].strip()
        end_str = parts[1].strip()
        if not start_str or not end_str:
            _log(f'[!] 跳过无效IP范围（起止IP为空）: {ip_str}')
            return result

        start_ip = ipaddress.ip_address(start_str)

        # 处理简写格式如 172.16.70.226-230
        end_parts = end_str.split('.')
        if len(end_parts) == 1:
            base_parts = start_str.split('.')
            if len(base_parts) == 4:
                end_str = f'{base_parts[0]}.{base_parts[1]}.{base_parts[2]}.{end_parts[0]}'
            else:
                _log(f'[!] 跳过无效IP范围（简写格式无法推断前三段）: {ip_str}')
                return result

        end_ip = ipaddress.ip_address(end_str)

        start_int = int(start_ip)
        end_int = int(end_ip)
        if start_int > end_int:
            # 自动交换，兼容 end-start 书写
            start_int, end_int = end_int, start_int

        span = end_int - start_int + 1
        # 大范围保护：超过 65536 (B类整段) 截断并警告，避免 OOM
        MAX_IP_RANGE = 65536
        if span > MAX_IP_RANGE:
            _log(f'[!] IP范围过大 ({span} 个地址)，截断为前 {MAX_IP_RANGE} 个: {ip_str}')
            end_int = start_int + MAX_IP_RANGE - 1
            span = MAX_IP_RANGE

        for ip_int in range(start_int, end_int + 1):
            result.append(str(ipaddress.ip_address(ip_int)))

    except ValueError as e:
        _log(f'[!] 跳过无效IP: {ip_str} ({e})')
    except Exception as e:
        _log(f'[!] 解析IP异常: {ip_str} ({e})')

    return result

def load_terminal_ip_table():
    global TERMINAL_IP_TABLE
    if TERMINAL_IP_TABLE is not None:
        return TERMINAL_IP_TABLE
    TERMINAL_IP_TABLE = {}
    # 优先使用外部导入路径，其次从文件系统查找
    ip_table_path = TERMINAL_IP_TABLE_PATH or _find_file("终端ip地址表.xlsx")
    if not ip_table_path or not os.path.exists(ip_table_path):
        return TERMINAL_IP_TABLE
    try:
        import pandas as pd
        df = pd.read_excel(ip_table_path)
        skipped = 0
        for _, row in df.iterrows():
            ip_value = row.iloc[2]
            dept_value = row.iloc[1] if len(row) > 1 else ''
            if pd.isna(ip_value):
                continue
            ip_str = str(ip_value).strip()
            dept_str = str(dept_value).strip() if not pd.isna(dept_value) else ''
            ips = _parse_ip_range(ip_str)
            if not ips:
                skipped += 1
                continue
            for ip in ips:
                TERMINAL_IP_TABLE[ip] = dept_str
        if skipped:
            _log(f'[*] 终端IP表: 跳过 {skipped} 行无效记录')
    except Exception as e:
        _log(f'[!] 加载终端IP表异常: {e}')
    return TERMINAL_IP_TABLE

def get_terminal_location(ip):
    ip = str(ip).strip()
    table = load_terminal_ip_table()
    return table.get(ip, '')

def is_valid_public_ip(ip):
    try:
        addr = ipaddress.ip_address(ip)
        return not (addr.is_private or addr.is_loopback or addr.is_reserved or addr.is_link_local or addr.is_multicast)
    except ValueError:
        return False

def local_ip_label(ip):
    try:
        ipaddress.ip_address(ip)
    except ValueError:
        return "无效IP"
    if not is_valid_public_ip(ip):
        return "内网/保留地址"
    return None

def format_online_result(data):
    if data.get("status") != "success":
        return data.get("message") or "查询失败"
    parts = []
    for key in ("country", "regionName", "city", "isp"):
        val = (data.get(key) or "").strip()
        if val:
            parts.append(val)
    return " ".join(parts) if parts else "未知"

BATCH_SIZE = 100
BATCH_INTERVAL = 1.5
XDB_FILE = None
DOWNLOAD_SOURCES = [
    "https://edgeone.gh-proxy.com/https://github.com/lionsoul2014/ip2region/raw/refs/heads/master/data/ip2region_v4.xdb",
    "https://gcore.jsdelivr.net/gh/lionsoul2014/ip2region/data/ip2region_v4.xdb",
    "https://raw.githubusercontent.com/lionsoul2014/ip2region/master/data/ip2region_v4.xdb",
    "https://gitee.com/lionsoul/ip2region/raw/master/data/ip2region_v4.xdb",
]

def download_xdb():
    for idx, url in enumerate(DOWNLOAD_SOURCES, 1):
        try:
            _log(f"[{idx}/{len(DOWNLOAD_SOURCES)}] 正在下载离线 IP 库: {url}")
            resp = safe_get(
                url, timeout=(30, 600), stream=True,
                ssl_fallback_msg="  SSL验证失败，跳过证书验证重试...",
            )
            total = int(resp.headers.get('content-length', 0))
            if total > 0:
                _set_progress(0, total)
            downloaded = 0
            # 确保目标目录存在（exe所在目录可能需要创建）
            os.makedirs(os.path.dirname(os.path.abspath(XDB_FILE)), exist_ok=True)
            with open(XDB_FILE, "wb") as f:
                for chunk in resp.iter_content(chunk_size=1024 * 1024):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total > 0:
                            _set_progress(downloaded, total)
            size_kb = os.path.getsize(XDB_FILE) // 1024
            if size_kb < 9000:
                _log(f"文件大小异常({size_kb}KB)，尝试下一个地址")
                try:
                    os.remove(XDB_FILE)
                except OSError:
                    pass
                continue
            _log(f"已保存: {XDB_FILE} ({size_kb} KB)")
            _set_progress(0, 0)
            return True
        except Exception as exc:
            _log(f"下载失败: {exc}")
            _set_progress(0, 0)
    _log("所有下载地址均失败，将使用在线 API 查询")
    _set_progress(0, 0)
    return False

def get_offline_searcher():
    global XDB_FILE
    if XDB_FILE is None:
        # 优先使用exe所在目录（打包后）
        if getattr(sys, 'frozen', False):
            XDB_FILE = os.path.join(os.path.dirname(sys.executable), "ip2region_v4.xdb")
        else:
            XDB_FILE = os.path.join(script_dir, "ip2region_v4.xdb")
    _log(f"[?] 检查本地IP库文件: {XDB_FILE}")
    # 检查文件是否存在及是否超过7天需要更新
    need_download = False
    if not os.path.exists(XDB_FILE):
        _log("[!] 本地IP库不存在，开始下载...")
        need_download = True
    else:
        file_mtime = os.path.getmtime(XDB_FILE)
        file_age_days = (time.time() - file_mtime) / 86400
        if file_age_days >= 7:
            _log(f"[!] 离线IP库已超过{int(file_age_days)}天，自动更新...")
            need_download = True
    if need_download:
        if not download_xdb():
            return None
    if not os.path.exists(XDB_FILE):
        _log("[!] 离线IP库文件不存在，将使用在线API查询")
        return None
    try:
        import ip2region.searcher as xdb
        import ip2region.util as util
        searcher = xdb.new_with_file_only(util.IPv4, str(XDB_FILE))
        _log("[+] 离线IP库加载成功")
        return searcher
    except Exception as exc:
        _log(f"[!] 离线IP库加载失败: {exc}，将使用在线API查询")
        return None

def parse_region(raw):
    parts = (raw or "").split("|")
    while len(parts) < 5:
        parts.append("")
    country, province, city, isp, _code = parts[:5]
    display = []
    if country and country not in ("0", "内网IP"):
        if country != "中国":
            display.append(country)
    if province and province not in ("0", ""):
        display.append(province)
    if city and city not in ("0", ""):
        display.append(city)
    if isp and isp not in ("0", ""):
        display.append(isp)
    location = " ".join(display) if display else (raw or "未知")
    stat_province = province if province not in ("0", "") else country
    stat_city = city if city not in ("0", "") else stat_province
    if not stat_province:
        stat_province = "未知"
    if not stat_city:
        stat_city = stat_province
    return location, stat_province, stat_city

def query_offline(searcher, ips):
    results = {}
    for ip in ips:
        if not is_valid_public_ip(ip):
            continue
        try:
            raw = searcher.search(ip)
            if not raw:
                results[ip] = ("未知", "未知", "未知")
            else:
                results[ip] = parse_region(raw)
        except Exception as exc:
            results[ip] = (f"查询失败({exc})", "查询失败", "查询失败")
    return results

def query_online_batch(ips):
    import requests
    results = {}
    public_ips = [ip for ip in ips if is_valid_public_ip(ip)]
    if not public_ips:
        return results
    url = "http://ip-api.com/batch?lang=zh-CN"
    fields = "status,message,country,regionName,city,isp,query"
    for i in range(0, len(public_ips), BATCH_SIZE):
        chunk = public_ips[i: i + BATCH_SIZE]
        payload = [{"query": ip, "fields": fields} for ip in chunk]
        try:
            resp = requests.post(url, json=payload, timeout=30, verify=get_requests_verify())
            resp.raise_for_status()
            for item in resp.json():
                ip_addr = item.get("query", "")
                results[ip_addr] = format_online_result(item)
        except Exception as e:
            _log(f"在线查询失败: {e}")
        if i + BATCH_SIZE < len(public_ips):
            time.sleep(BATCH_INTERVAL)
    return results

def query_all_ips(ips):
    results = {}
    for ip in ips:
        local = local_ip_label(ip)
        if local:
            results[ip] = (local, local, local)
    pending = [ip for ip in ips if ip not in results]
    if not pending:
        return results
    searcher = get_offline_searcher()
    if searcher:
        _log(f"使用本地 ip2region 离线库查询 {len(pending)} 个IP...")
        results.update(query_offline(searcher, pending))
        found = sum(1 for v in results.values() if v[0] not in ("未知", "查询失败"))
        _log(f"离线查询完成，成功 {found}/{len(pending)}")
        return results
    _log(f"使用在线 API 批量查询 {len(pending)} 个IP...")
    online = query_online_batch(pending)
    for ip in pending:
        location = online.get(ip, "未知")
        parts = location.split()
        stat_province = parts[1] if len(parts) >= 2 and parts[0] == "中国" else parts[0] if parts else "未知"
        stat_city = parts[2] if len(parts) >= 3 else stat_province
        results[ip] = (location, stat_province, stat_city)
    return results

def extract_source_ips(file_path):
    import pandas as pd
    df = pd.read_excel(file_path)
    df.columns = df.columns.str.strip()
    col_names = list(df.columns)
    src_ip_col = None
    dst_ip_col = None
    for col in col_names:
        col_lower = col.lower() if isinstance(col, str) else ''
        if '源' in col and 'ip' in col_lower:
            src_ip_col = col
        elif '目' in col and 'ip' in col_lower:
            dst_ip_col = col
    if not src_ip_col or not dst_ip_col:
        _log(f"[!] 无法找到必需的列，文件: {file_path}")
        return [], [], []
    df = df.rename(columns={src_ip_col: '源 IP', dst_ip_col: '目标 IP'})
    # 分离已排除IP（本地公网IP）和待分析IP
    excluded_mask = df['源 IP'].apply(is_excluded_ip)
    excluded_df = df[excluded_mask].copy()
    df = df[~excluded_mask]
    # 收集被排除的源IP（去重）
    excluded_ips = excluded_df['源 IP'].drop_duplicates().tolist() if len(excluded_df) > 0 else []
    def get_ip_type(ip_str):
        try:
            ip = ipaddress.ip_address(str(ip_str).strip())
            return 'internal' if ip.is_private else 'external'
        except ValueError:
            return 'invalid'
    df['源IP类型'] = df['源 IP'].apply(get_ip_type)
    df['目标IP类型'] = df['目标 IP'].apply(get_ip_type)
    external_to_internal = df[(df['源IP类型'] == 'external') & (df['目标IP类型'] == 'internal')].copy()
    internal_df = df[df['源IP类型'] == 'internal'].copy()
    external_ips = []
    if len(external_to_internal) > 0:
        external_ips = external_to_internal['源 IP'].drop_duplicates().tolist()
    internal_ips = []
    if len(internal_df) > 0:
        internal_ips = internal_df['源 IP'].drop_duplicates().tolist()
    return external_ips, internal_ips, excluded_ips

def load_config(files=None, local_geos=None):
    """读取配置：config.ini 为兜底 + 从文件自动获取（解放 config 硬编码）。

    files：本次生成的安全告警文件列表（Path）。提供时：
      - internal_zones ← 告警"源区域"列自动提取（config 值合并兜底）
      - local_geos    ← 告警"源地理信息"列提取"长春"等（config 值合并兜底）
    local_geos：GUI 自定义输入（逗号分隔，如 "长春,上海"），与自动提取值合并。
    excluded_ips ← 业务ip.xlsx 自动加载（模块级 EXCLUDED_IP_NETWORKS）。
    probes       ← 业务ip.xlsx[探针]sheet 优先，否则 config [health] probes。
    """
    # 排除IP：业务ip.xlsx 自动加载（幂等，config 的 excluded_ips 已不再需要）
    _auto_load_excluded_ips()

    cfg = configparser.ConfigParser()
    # 优先从exe目录读取，其次从临时解压目录；缺失时全部回退默认值（不崩溃）
    config_path = _find_file('config.ini')
    if not config_path or not os.path.exists(config_path):
        _log('[!] config.ini 不存在，全部使用默认配置')
    else:
        try:
            cfg.read(config_path, encoding='utf-8')
        except Exception as e:
            _log(f'[!] 读取 config.ini 失败: {e}，全部使用默认配置')

    def _get(section, option, fallback=''):
        """缺段/缺键均回退默认值，避免 KeyError / NoSectionError"""
        if cfg.has_option(section, option):
            return cfg.get(section, option)
        return fallback

    def _getint(section, option, fallback=0):
        if cfg.has_option(section, option):
            try:
                return cfg.getint(section, option)
            except ValueError:
                _log(f'[!] config.ini [{section}] {option} 不是整数，使用默认 {fallback}')
        return fallback

    conf = {}
    conf['title'] = _get('base', 'report_title', '网络安全值守保障日报')
    conf['pattern'] = _get('base', 'input_pattern', '*.xlsx')
    conf['out_dir'] = _get('base', 'output_dir', 'output')
    conf['intel_file'] = _get('base', 'intel_file', 'intel.csv')
    ranges_raw = _get('network', 'ranges', '')
    nets = [ipaddress.ip_network(x.strip()) for x in ranges_raw.splitlines()
            if x.strip() and not x.strip().startswith('#')]
    terminal_ips = set()
    # 复用 load_terminal_ip_table()（已缓存，避免重复读取同一文件）
    terminal_table = load_terminal_ip_table()
    if terminal_table:
        terminal_ips = set(terminal_table.keys())
        terminal_nets = list(ipaddress.collapse_addresses([ipaddress.ip_address(ip) for ip in terminal_ips]))
        nets.extend(terminal_nets)
        _log(f'[+] 加载终端IP地址表: {len(terminal_ips)}个IP, 合并为{len(terminal_nets)}个网段')
    conf['nets'] = nets
    conf['zones'] = {x.strip() for x in _get('network', 'internal_zones', '').split(',') if x.strip()}
    conf['geos'] = {x.strip() for x in _get('network', 'local_geos', '').split(',') if x.strip()}
    probes_raw = _get('health', 'probes', '')
    conf['probes'] = [tuple(x.strip().split('|')) for x in probes_raw.splitlines()
                      if x.strip() and not x.strip().startswith('#') and '|' in x]
    conf['retention'] = _getint('health', 'log_retention_days', 180)
    conf['top'] = _getint('report', 'top_events', 5)
    conf['crit_levels'] = {x.strip() for x in _get('report', 'critical_levels', '严重,高危').split(',') if x.strip()}
    conf['ban_levels'] = {x.strip() for x in _get('report', 'ban_levels', '高危,严重').split(',') if x.strip()}

    # ---- 从文件自动增强（config 留空时生效，有值时合并）----
    if files:
        alert_zones = extract_zones_from_alerts(files)
        if alert_zones:
            conf['zones'] |= alert_zones
        alert_geos = extract_geos_from_alerts(files)
        if alert_geos:
            conf['geos'] |= alert_geos
    # GUI 自定义本地归属地关键词（逗号分隔）与自动提取合并
    if local_geos:
        custom = {x.strip() for x in str(local_geos).split(',') if x.strip()}
        if custom:
            conf['geos'] |= custom
    # probes：业务ip.xlsx[探针]sheet 优先（config [health] probes 作为兜底）
    excel_probes = load_probes_from_excel()
    if excel_probes:
        conf['probes'] = excel_probes
    return conf

def generate_ip_report(files, date, local_geos=None):
    all_external_ips = set()
    all_internal_ips = set()
    all_excluded_ips = set()
    for f in files:
        _log(f'[+] 处理文件: {f.name}')
        file_path = str(f)
        external_ips, internal_ips, excluded_ips = extract_source_ips(file_path)
        all_external_ips.update(external_ips)
        all_internal_ips.update(internal_ips)
        all_excluded_ips.update(excluded_ips)
    _log(f'[+] 外网攻击IP去重后共 {len(all_external_ips)} 个，开始查询归属地...')
    _log(f'[+] 内网IP去重后共 {len(all_internal_ips)} 个，开始查询归属地...')
    _log(f'[+] 本地公网IP(已排除)去重后共 {len(all_excluded_ips)} 个')
    external_ip_list = list(all_external_ips)
    internal_ip_list = list(all_internal_ips)
    excluded_ip_list = list(all_excluded_ips)
    location_map = query_all_ips(external_ip_list)
    # 查询排除IP的归属地
    excluded_location_map = query_all_ips(excluded_ip_list) if excluded_ip_list else {}
    load_terminal_ip_table()
    # 加载 local_geos 配置用于标记本地IP（传入 files 自动从告警提取区域/归属地；local_geos 为 GUI 自定义）
    try:
        conf = load_config(files, local_geos=local_geos)
        local_geos_set = conf.get('geos', set())
    except Exception as e:
        _log(f'[!] 加载配置失败，local_geos 回退为空: {e}')
        local_geos_set = set()
    # 使用 runtime_dir（脚本/exe所在目录）作为输出基础目录
    out_dir = Path(runtime_dir) / 'output'
    out_dir.mkdir(exist_ok=True)
    output_file = out_dir / f'IP归属分析结果-{date}.xlsx'

    # 懒加载openpyxl模块
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

    wb = Workbook()
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    # 本地IP高亮填充
    local_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    ws1 = wb.active
    ws1.title = '外网攻击IP归属'
    headers1 = ['序号', 'IP地址', '归属地', '备注']
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    for idx, ip in enumerate(external_ip_list, start=1):
        location, _, _ = location_map.get(ip, ("未知", "未知", "未知"))
        # 检查是否为本地IP
        remark = ''
        is_local = False
        if local_geos_set and any(g in str(location) for g in local_geos_set):
            remark = '本地IP'
            is_local = True
        ws1.append([idx, ip, location, remark])
        for col in range(1, 5):
            ws1.cell(row=idx + 1, column=col).border = thin_border
            if is_local:
                ws1.cell(row=idx + 1, column=col).fill = local_fill
        ws1.cell(row=idx + 1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=idx + 1, column=2).alignment = Alignment(horizontal='left', vertical='center')
        ws1.cell(row=idx + 1, column=3).alignment = Alignment(wrap_text=True, vertical='center')
        ws1.cell(row=idx + 1, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 40
    ws1.column_dimensions["D"].width = 12
    ws2 = wb.create_sheet('内网IP归属')
    headers2 = ['序号', 'IP地址', '归属地']
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    for idx, ip in enumerate(internal_ip_list, start=1):
        location = get_terminal_location(ip)
        ws2.append([idx, ip, location])
        for col in range(1, 4):
            ws2.cell(row=idx + 1, column=col).border = thin_border
        ws2.cell(row=idx + 1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=idx + 1, column=2).alignment = Alignment(horizontal='left', vertical='center')
        ws2.cell(row=idx + 1, column=3).alignment = Alignment(wrap_text=True, vertical='center')
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 40
    # 第三 sheet：本地公网IP（已排除）
    if excluded_ip_list:
        ws3 = wb.create_sheet('本地公网IP(已排除)')
        headers3 = ['序号', 'IP地址', '归属地', '类型']
        ws3.append(headers3)
        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        for idx, ip in enumerate(excluded_ip_list, start=1):
            location, _, _ = excluded_location_map.get(ip, ("未知", "未知", "未知"))
            label = EXCLUDED_IP_LABELS.get(ip, '业务IP')
            ws3.append([idx, ip, location, label])
            for col in range(1, 5):
                ws3.cell(row=idx + 1, column=col).border = thin_border
                ws3.cell(row=idx + 1, column=col).fill = local_fill
            ws3.cell(row=idx + 1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws3.cell(row=idx + 1, column=2).alignment = Alignment(horizontal='left', vertical='center')
            ws3.cell(row=idx + 1, column=3).alignment = Alignment(wrap_text=True, vertical='center')
            ws3.cell(row=idx + 1, column=4).alignment = Alignment(horizontal='center', vertical='center')
        ws3.column_dimensions["A"].width = 8
        ws3.column_dimensions["B"].width = 18
        ws3.column_dimensions["C"].width = 40
        ws3.column_dimensions["D"].width = 22
    try:
        wb.save(output_file)
    except PermissionError:
        # 文件被占用（可能在Excel中打开），使用带时间戳的备用文件名
        alt_file = out_dir / f'IP归属分析结果-{date}-{datetime.now().strftime("%H%M%S")}.xlsx'
        wb.save(alt_file)
        _log(f'[!] 原文件被占用，已保存为: {alt_file}')
        output_file = alt_file
    _log(f'[✓] IP归属分析完成，结果已保存至: {output_file}')
    return output_file
