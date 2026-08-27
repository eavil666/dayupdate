"""openpyxl 内嵌 Excel 图表演示：基于真实告警数据生成 demo_chart.xlsx
三种图表：威胁等级柱状图 / 攻击类型水平条形图 / 告警趋势折线图（按小时）
"""

import glob

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = r"E:\script\python\日报update\demo_chart.xlsx"

# ---------- 1. 统计真实告警数据 ----------
files = sorted(glob.glob(r"E:\script\python\日报update\安全告警*.xlsx"))
dfs = []
for f in files:
    d = pd.read_excel(f)
    dfs.append(d)
df = pd.concat(dfs, ignore_index=True)

# 威胁等级分布
level_counts = df["威胁等级"].value_counts()
# 攻击类型 TOP6
att_counts = df["攻击类型"].value_counts().head(6)
# 按小时趋势（用"最近发生时间"列，取前 8 位时间）
time_col = None
for c in ("最近发生时间", "发生时间", "时间", "时间戳"):
    if c in df.columns:
        time_col = c
        break
hour_counts = pd.Series(dtype=int)
if time_col:
    try:
        hours = pd.to_datetime(df[time_col], errors="coerce").dt.hour
        hour_counts = hours.value_counts().sort_index()
    except Exception:
        pass

# ---------- 2. 构建工作簿 ----------
wb = Workbook()
ws = wb.active
ws.title = "图表数据"

# 样式
hdr_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="4472C4")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
title_font = Font(name="微软雅黑", bold=True, size=12)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


# 数据块1：威胁等级分布（B2 起，A 列放标题）
ws["A1"] = "威胁等级分布"
ws["A1"].font = title_font
ws.append(["等级", "数量"])
style_header(ws, 2, 2)
for lv, cnt in level_counts.items():
    ws.append([str(lv), int(cnt)])
n_level_rows = len(level_counts)
level_last = 2 + n_level_rows

# 数据块2：攻击类型 TOP6
ws.cell(row=level_last + 2, column=1, value="攻击类型 TOP6")
ws.cell(row=level_last + 2, column=1).font = title_font
ws.cell(row=level_last + 3, column=1, value="类型")
ws.cell(row=level_last + 3, column=2, value="数量")
style_header(ws, level_last + 3, 2)
for name, cnt in att_counts.items():
    ws.cell(row=level_last + 3 + 1 + len(ws.cell(row=level_last + 4).value if False else []), column=1)
for i, (name, cnt) in enumerate(att_counts.items()):
    ws.cell(row=level_last + 4 + i, column=1, value=str(name))
    ws.cell(row=level_last + 4 + i, column=2, value=int(cnt))
att_last = level_last + 3 + len(att_counts)

# 数据块3：按小时趋势
if not hour_counts.empty:
    ws.cell(row=att_last + 2, column=1, value="告警按小时趋势")
    ws.cell(row=att_last + 2, column=1).font = title_font
    ws.cell(row=att_last + 3, column=1, value="小时")
    ws.cell(row=att_last + 3, column=2, value="告警数")
    style_header(ws, att_last + 3, 2)
    for h, cnt in hour_counts.items():
        ws.cell(row=att_last + 4 + h, column=1, value=f"{int(h)}时")
        ws.cell(row=att_last + 4 + h, column=2, value=int(cnt))
    trend_last = att_last + 3 + len(hour_counts)
else:
    trend_last = att_last

# 加边框
for row in ws.iter_rows(min_row=2, max_row=trend_last, min_col=1, max_col=2):
    for cell in row:
        cell.border = border

# 列宽
ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 12

# ---------- 3. 嵌入图表 ----------
# 图表1：威胁等级柱状图（放 C2）
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "威胁等级分布"
chart1.y_axis.title = "数量"
data1 = Reference(ws, min_col=2, min_row=2, max_row=level_last)
cats1 = Reference(ws, min_col=1, min_row=3, max_row=level_last)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.dataLabels = DataLabelList()
chart1.dataLabels.showVal = True
chart1.width = 14
chart1.height = 8
ws.add_chart(chart1, "C2")

# 图表2：攻击类型水平条形图（放 C22）
chart2 = BarChart()
chart2.type = "bar"
chart2.style = 11
chart2.title = "攻击类型 TOP6"
chart2.y_axis.title = "类型"
data2 = Reference(ws, min_col=2, min_row=level_last + 3, max_row=att_last)
cats2 = Reference(ws, min_col=1, min_row=level_last + 4, max_row=att_last)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.dataLabels = DataLabelList()
chart2.dataLabels.showVal = True
chart2.width = 14
chart2.height = 8
ws.add_chart(chart2, "C22")

# 图表3：按小时折线图（有数据才加）
if not hour_counts.empty:
    chart3 = LineChart()
    chart3.style = 12
    chart3.title = "告警按小时趋势"
    chart3.y_axis.title = "告警数"
    chart3.x_axis.title = "小时"
    data3 = Reference(ws, min_col=2, min_row=att_last + 3, max_row=trend_last)
    cats3 = Reference(ws, min_col=1, min_row=att_last + 4, max_row=trend_last)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.width = 20
    chart3.height = 8
    ws.add_chart(chart3, "F22")

wb.save(OUT)
print("已生成:", OUT)
print(f"  图表1 威胁等级柱状图: {n_level_rows} 项")
print(f"  图表2 攻击类型条形图: {len(att_counts)} 项")
if not hour_counts.empty:
    print(f"  图表3 按小时折线图: {len(hour_counts)} 个时点")
else:
    print("  图表3: 无时间列，跳过")
