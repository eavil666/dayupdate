"""main.py 业务纯逻辑测试（不触网、不读写真实业务文件）"""
import ipaddress

import main


def test_is_private_ip():
    assert main.is_private_ip('192.168.1.1') is True
    assert main.is_private_ip('10.0.0.1') is True
    assert main.is_private_ip('8.8.8.8') is False


def test_is_valid_public_ip():
    assert main.is_valid_public_ip('8.8.8.8') is True
    assert main.is_valid_public_ip('192.168.1.1') is False   # 私网
    assert main.is_valid_public_ip('999.1.1.1') is False     # 非法


def test_local_ip_label():
    assert main.local_ip_label('abc') == '无效IP'
    assert main.local_ip_label('192.168.1.1') == '内网/保留地址'
    assert main.local_ip_label('8.8.8.8') is None


def test_parse_region():
    # ip2region 格式：国家|省份|城市|ISP|编码
    loc, prov, city = main.parse_region('中国|广东省|广州市|电信|0')
    assert loc == '广东省 广州市 电信'
    assert prov == '广东省' and city == '广州市'
    # 字段不全时回退
    loc2, prov2, city2 = main.parse_region('中国|0')
    assert loc2 == '中国|0' and prov2 == '中国' and city2 == '中国'
    assert main.parse_region(None) == ('未知', '未知', '未知')


def test_format_online_result():
    assert main.format_online_result({'status': 'success',
                                      'country': '中国', 'regionName': '广东',
                                      'city': '广州', 'isp': '电信'}) == '中国 广东 广州 电信'
    assert main.format_online_result({'status': 'fail', 'message': 'private range'}) == 'private range'
    assert main.format_online_result({'status': 'fail'}) == '查询失败'


def _conf():
    return {
        'nets': [ipaddress.ip_network('10.0.0.0/8')],
        'zones': {'内网'},
        'geos': {'北京'},
    }


def test_classify():
    c = _conf()
    assert main.classify('10.1.1.1', '', '', c) == '内网'
    assert main.classify('172.16.0.5', '内网', '北京', c) == '内网'
    assert main.classify('8.8.8.8', '', '', c) == '外网'
    assert main.classify('172.16.0.5', '', '', c) == '待确认'
    assert main.classify('bad-ip', '', '', c) == '未知'


def test_parse_ip_range():
    r = main._parse_ip_range('10.0.0.1-10.0.0.10')
    assert len(r) == 10
    assert r[0] == '10.0.0.1' and r[-1] == '10.0.0.10'
    # 简写范围
    r2 = main._parse_ip_range('172.16.70.226-230')
    assert len(r2) == 5 and r2[0] == '172.16.70.226' and r2[-1] == '172.16.70.230'
    # 单 IP
    assert main._parse_ip_range('8.8.8.8') == ['8.8.8.8']
    # 不支持 CIDR / 空串 → 空列表
    assert main._parse_ip_range('192.168.1.0/24') == []
    assert main._parse_ip_range('') == []


def test_is_excluded_ip(monkeypatch):
    # 拆分后 is_excluded_ip 读取的是 ipdb 模块的全局状态，须 patch ipdb
    import ipdb
    monkeypatch.setattr(ipdb, 'EXCLUDED_IP_NETWORKS',
                        [ipaddress.ip_network('10.0.0.0/8'),
                         ipaddress.ip_address('192.168.9.9')])
    assert main.is_excluded_ip('10.1.2.3') is True
    assert main.is_excluded_ip('192.168.9.9') is True
    assert main.is_excluded_ip('8.8.8.8') is False
    assert main.is_excluded_ip('bad') is False


def test_load_config_missing_sections(tmp_path, monkeypatch):
    """config.ini 缺失/缺段/非法整数时回退默认值，不崩溃（P0 容错加固）"""
    import ipdb
    monkeypatch.setattr(ipdb, 'load_terminal_ip_table', lambda: {})

    # 1) config.ini 完全不存在
    monkeypatch.setattr(ipdb, '_find_file', lambda name: str(tmp_path / 'nope.ini'))
    conf = ipdb.load_config()
    assert conf['title'] == '网络安全值守保障日报'
    assert conf['pattern'] == '*.xlsx'
    assert conf['out_dir'] == 'output'
    assert conf['intel_file'] == 'intel.csv'
    assert conf['nets'] == []
    assert conf['probes'] == []
    assert conf['retention'] == 180
    assert conf['top'] == 5
    assert conf['crit_levels'] == {'严重', '高危'}
    assert conf['ban_levels'] == {'高危', '严重'}

    # 2) 残缺 config（只有 [base] 一段）
    broken = tmp_path / 'broken.ini'
    broken.write_text('[base]\nreport_title = 测试日报\n', encoding='utf-8')
    monkeypatch.setattr(ipdb, '_find_file', lambda name: str(broken))
    conf2 = ipdb.load_config()
    assert conf2['title'] == '测试日报'       # base 段生效
    assert conf2['retention'] == 180          # health 段缺失回退
    assert conf2['top'] == 5                  # report 段缺失回退
    assert conf2['probes'] == []              # health 段缺失回退

    # 3) 非法整数（log_retention_days = abc）回退默认
    bad = tmp_path / 'bad.ini'
    bad.write_text('[health]\nlog_retention_days = abc\n', encoding='utf-8')
    monkeypatch.setattr(ipdb, '_find_file', lambda name: str(bad))
    conf3 = ipdb.load_config()
    assert conf3['retention'] == 180


# ---------- 配置解放：从 Excel/告警文件自动获取 ----------

def test_extract_zones_geos_from_alerts(tmp_path):
    """从告警文件源区域/源地理信息列提取（排除默认区域噪声）"""
    import pandas as pd

    import ipdb
    alert = tmp_path / 'alerts_test.xlsx'
    pd.DataFrame({
        '源区域': ['集团四楼', '默认区域', '后楼二楼西', '集团四楼'],
        '源地理信息': ['吉林-长春', '吉林-长春', '美国', '中国'],
    }).to_excel(alert, index=False)
    zones = ipdb.extract_zones_from_alerts([alert])
    assert zones == {'集团四楼', '后楼二楼西'}   # 默认区域被排除
    geos = ipdb.extract_geos_from_alerts([alert])
    assert geos == {'长春'}                       # 吉林-长春 → 长春


def test_load_probes_from_excel(tmp_path, monkeypatch):
    """业务ip.xlsx 带"探针"sheet 时读取探针；无 sheet 返回空"""
    from openpyxl import Workbook

    import ipdb
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['ip', '说明'])
    ws.append(['1.2.3.4', '测试'])
    ws2 = wb.create_sheet('探针')
    ws2.append(['名称', 'IP地址'])
    ws2.append(['探针A', '172.16.1.1'])
    ws2.append(['探针B', '172.16.1.2'])
    p = tmp_path / 'biz.xlsx'
    wb.save(p)
    monkeypatch.setattr(ipdb, '_find_file', lambda name: str(p))
    rows = ipdb.load_probes_from_excel()
    assert rows == [('探针A', '172.16.1.1'), ('探针B', '172.16.1.2')]
    # 无探针 sheet
    wb3 = Workbook()
    wb3.active.append(['ip', '说明'])
    p3 = tmp_path / 'biz3.xlsx'
    wb3.save(p3)
    monkeypatch.setattr(ipdb, '_find_file', lambda name: str(p3))
    assert ipdb.load_probes_from_excel() == []


def test_load_config_enhance_from_alerts(tmp_path, monkeypatch):
    """load_config(files)：告警提取的 zones/geos 合并进 conf（config 留空时生效）"""
    import ipdb
    # 隔离文件依赖
    monkeypatch.setattr(ipdb, '_auto_load_excluded_ips', lambda: None)
    monkeypatch.setattr(ipdb, 'load_terminal_ip_table', lambda: {})
    monkeypatch.setattr(ipdb, 'load_probes_from_excel', lambda: [])
    monkeypatch.setattr(ipdb, '_find_file', lambda name: str(tmp_path / 'none.ini'))
    monkeypatch.setattr(ipdb, 'extract_zones_from_alerts', lambda files: {'集团四楼', '一卡通'})
    monkeypatch.setattr(ipdb, 'extract_geos_from_alerts', lambda files: {'长春'})
    conf = ipdb.load_config(files=[tmp_path / 'a.xlsx'])
    assert '集团四楼' in conf['zones'] and '一卡通' in conf['zones']
    assert '长春' in conf['geos']
    assert conf['retention'] == 180       # 其余默认值不受影响
    assert conf['probes'] == []           # 探针 sheet 为空


def test_load_config_custom_geos(tmp_path, monkeypatch):
    """GUI 自定义 local_geos（逗号分隔）与 config 兜底合并"""
    import ipdb
    monkeypatch.setattr(ipdb, '_auto_load_excluded_ips', lambda: None)
    monkeypatch.setattr(ipdb, 'load_terminal_ip_table', lambda: {})
    monkeypatch.setattr(ipdb, 'load_probes_from_excel', lambda: [])
    # config 缺失 → 无兜底关键词
    monkeypatch.setattr(ipdb, '_find_file', lambda name: str(tmp_path / 'none.ini'))
    conf = ipdb.load_config(files=[], local_geos='长春,上海')
    assert conf['geos'] == {'长春', '上海'}
    # 空/空白输入 → 无新增
    conf2 = ipdb.load_config(files=[], local_geos='  ')
    assert conf2['geos'] == set()
