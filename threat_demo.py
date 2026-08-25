# -*- coding: utf-8 -*-
"""威胁分级接入演示：提取告警外网源 IP -> 威胁分级 -> 生成带分级列的 Excel"""
import glob
import os
import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import threat_check
import ipdb

OUT = r'E:\script\python\日报update\demo_threat.xlsx'
CACHE = r'E:\script\python\日报update\threat_feeds_cache.json'

# ---------- 1. 提取告警中的外网源 IP ----------
files = sorted(glob.glob(r'E:\script\python\日报update\安全告警*.xlsx'))
ext_ips = set()
for f in files:
    ext, internal, excluded = ipdb.extract_source_ips(f)
    ext_ips.update(ext)
print(f'告警文件 {len(files)} 个，外网源 IP 去重后 {len(ext_ips)} 个')

# ---------- 2. 加载威胁名单（带磁盘缓存） ----------
bad_ips, sources = threat_check.load_bad_ips(CACHE)
print(f'威胁名单: {len(bad_ips)} 条')

# ---------- 3. 批量分级 ----------
levels = {'Critical': 0, 'High': 0, 'Clean': 0, '未查': 0}
rows = []
for ip in sorted(ext_ips):
    if ip in bad_ips:
        level, detail = threat_check.check_ip(ip, bad_ips, sources)
    else:
        level, detail = 'Clean', []
    levels[level if level in levels else '未查'] = levels.get(level if level in levels else '未查', 0) + 1
    rows.append((ip, level, ';'.join(detail)))

# ---------- 4. 归属地（直接读本地 xdb，跳过 7 天自动更新） ----------
geo = {}
try:
    import ip2region.searcher as xdb
    import ip2region.util as util
    s = xdb.new_with_file_only(util.IPv4, r'E:\script\python\日报update\ip2region_v4.xdb')
    for ip in [r[0] for r in rows]:
        try:
            res = s.search(ip)  # 返回 '国家|区域|省份|城市|ISP'
            geo[ip] = res.split('|') if res else ('', '', '')
        except Exception:
            geo[ip] = ('', '', '')
except Exception as e:
    print('归属查询失败（跳过）:', e)

# ---------- 5. 生成 Excel（外网 sheet + 分级列） ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = '外网IP威胁分级'
headers = ['序号', 'IP', '归属地', '威胁分级', '命中威胁源']
ws.append(headers)

# 表头样式
hdr_font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='4472C4')
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# 分级配色（中国惯例：红=危险）
level_fill = {
    'Critical': PatternFill('solid', fgColor='C00000'),
    'High':     PatternFill('solid', fgColor='FF6B6B'),
    'Clean':    PatternFill('solid', fgColor='C6EFCE'),
    '未查':     PatternFill('solid', fgColor='D9D9D9'),
}
level_font = {
    'Critical': Font(name='微软雅黑', bold=True, size=10, color='FFFFFF'),
    'High':     Font(name='微软雅黑', bold=True, size=10, color='FFFFFF'),
    'Clean':    Font(name='微软雅黑', size=10, color='006100'),
    '未查':     Font(name='微软雅黑', size=10, color='595959'),
}

for idx, (ip, level, detail) in enumerate(rows, 1):
    loc = geo.get(ip, ('', '', ''))[0] if geo else ''
    ws.append([idx, ip, loc, level, detail])
    r = ws.max_row
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        cell.font = Font(name='微软雅黑', size=10)
    # 分级列配色
    lv_cell = ws.cell(row=r, column=4)
    if level in level_fill:
        lv_cell.fill = level_fill[level]
        lv_cell.font = level_font[level]
    lv_cell.alignment = Alignment(horizontal='center', vertical='center')

# 列宽
for col, w in zip('ABCDE', [8, 18, 26, 12, 30]):
    ws.column_dimensions[col].width = w

wb.save(OUT)
print('\n已生成:', OUT)
print('分级统计:', levels)
print('\n命中示例（前10条）:')
hit = [r for r in rows if r[1] != 'Clean']
for ip, lv, dt in hit[:10]:
    print(f'  {ip:<18} {lv:<10} {dt}')
print(f'\n命中总数: {len(hit)} / {len(rows)}')
